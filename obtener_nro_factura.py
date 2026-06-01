from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Iterable

import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


# Encabezados exactos de la plantilla compatible con el sistema.
ENCABEZADOS = [
    "Nro_Referencia",
    "Valor original",
    "Saldo de la factura",
    "A pagar",
]

TIPO_FACTURA = "FAC"
TIPO_NOTA_CREDITO = "NCR"


@dataclass(frozen=True)
class DocumentoPDF:
    tipo: str  # FAC o NCR
    nro_original: str  # 0020010011951
    nro_formateado: str  # 002-001-0011951
    imp_factura: str  # 1480920 o 190801, sin separador de miles ni signo negativo


@dataclass(frozen=True)
class ResultadoProceso:
    ruta_facturas: Path
    ruta_notas_credito: Path
    documentos_pdf: list[DocumentoPDF]
    total_facturas_pdf: int
    total_notas_credito_pdf: int
    total_facturas_excel: int
    total_notas_credito_excel: int
    fecha_archivo: str


# Lee líneas como:
# 274287835 - FACTURA - 0020010011951 1.480.920 0
# 277342043 - NOTA CRED. - 0010010005107 -190.801 0
# Importante: captura SOLO el primer importe posterior al documento = Imp. Factura.
PATRON_LINEA_DOCUMENTO = re.compile(
    r"^\s*"
    r"(?P<nro_interno>\d+)\s+-\s+"
    r"(?P<tipo>FACTURA|NOTA\s+CRED\.?)\s+-\s+"
    r"(?P<nro_doc>\d{13})\s+"
    r"(?P<imp_factura>-?\d{1,3}(?:\.\d{3})*|-?\d+)"
    r"(?:\s+(?P<imp_dcto>-?\d{1,3}(?:\.\d{3})*|-?\d+))?"
    r"\s*$",
    re.IGNORECASE,
)


def obtener_fecha_archivo(fecha: date | None = None) -> str:
    """Devuelve la fecha para nombres de archivo con formato YYYYMMDD."""
    return (fecha or date.today()).strftime("%Y%m%d")


def formatear_nro_documento(nro_doc: str) -> str:
    """Convierte 0020010011951 en 002-001-0011951."""
    nro_limpio = re.sub(r"\D", "", str(nro_doc or ""))

    if len(nro_limpio) != 13:
        raise ValueError(f"El número de documento no tiene 13 dígitos: {nro_doc}")

    return f"{nro_limpio[:3]}-{nro_limpio[3:6]}-{nro_limpio[6:]}"


def convertir_imp_factura(valor: str, quitar_signo: bool = False) -> str:
    """
    Convierte Imp. Factura a texto numérico limpio.

    Ejemplos:
    - "1.480.920" -> "1480920"
    - "-190.801" -> "190801" si quitar_signo=True
    - "-190.801" -> "-190801" si quitar_signo=False

    No conserva separadores de miles.
    """
    valor_limpio = str(valor or "").strip().replace(".", "")

    if not re.fullmatch(r"-?\d+", valor_limpio):
        raise ValueError(f"Importe inválido en columna Imp. Factura: {valor}")

    if quitar_signo:
        valor_limpio = valor_limpio.lstrip("-")

    return valor_limpio


def extraer_documentos_pdf(ruta_pdf: str | Path) -> list[DocumentoPDF]:
    """
    Lee el PDF y extrae facturas/notas de crédito con su Imp. Factura.

    Devuelve documentos en el mismo orden en que aparecen en el PDF.
    Elimina duplicados exactos por tipo, documento e importe.
    Para notas de crédito, guarda el importe sin signo negativo.
    """
    ruta_pdf = Path(ruta_pdf)

    if not ruta_pdf.exists():
        raise FileNotFoundError(f"No existe el archivo PDF: {ruta_pdf}")

    documentos: list[DocumentoPDF] = []
    vistos: set[tuple[str, str, str]] = set()

    with pdfplumber.open(str(ruta_pdf)) as pdf:
        for pagina in pdf.pages:
            texto = pagina.extract_text() or ""

            for linea in texto.splitlines():
                match = PATRON_LINEA_DOCUMENTO.match(linea)
                if not match:
                    continue

                tipo_pdf = match.group("tipo").upper()
                tipo = TIPO_NOTA_CREDITO if "NOTA" in tipo_pdf else TIPO_FACTURA
                nro_original = match.group("nro_doc")
                nro_formateado = formatear_nro_documento(nro_original)
                imp_factura = convertir_imp_factura(
                    match.group("imp_factura"),
                    quitar_signo=(tipo == TIPO_NOTA_CREDITO),
                )

                clave = (tipo, nro_original, imp_factura)
                if clave in vistos:
                    continue

                vistos.add(clave)
                documentos.append(
                    DocumentoPDF(
                        tipo=tipo,
                        nro_original=nro_original,
                        nro_formateado=nro_formateado,
                        imp_factura=imp_factura,
                    )
                )

    return documentos


def separar_documentos(documentos: Iterable[DocumentoPDF]) -> tuple[list[DocumentoPDF], list[DocumentoPDF]]:
    """Separa documentos en facturas y notas de crédito."""
    facturas: list[DocumentoPDF] = []
    notas_credito: list[DocumentoPDF] = []

    for doc in documentos:
        if doc.tipo == TIPO_FACTURA:
            facturas.append(doc)
        elif doc.tipo == TIPO_NOTA_CREDITO:
            notas_credito.append(doc)

    return facturas, notas_credito


def contar_por_tipo(documentos: Iterable[DocumentoPDF]) -> dict[str, int]:
    """Retorna cantidades por tipo FAC/NCR."""
    conteo = {TIPO_FACTURA: 0, TIPO_NOTA_CREDITO: 0}
    for doc in documentos:
        if doc.tipo in conteo:
            conteo[doc.tipo] += 1
    return conteo


def crear_rutas_salida(ruta_pdf: str | Path, carpeta_salida: str | Path, fecha_archivo: str | None = None) -> tuple[Path, Path]:
    """Crea rutas de salida usando nombre del PDF, tipo y fecha YYYYMMDD."""
    ruta_pdf = Path(ruta_pdf)
    carpeta_salida = Path(carpeta_salida)

    if not carpeta_salida.exists() or not carpeta_salida.is_dir():
        raise FileNotFoundError(f"La carpeta de guardado no existe: {carpeta_salida}")

    fecha_archivo = fecha_archivo or obtener_fecha_archivo()
    nombre_base = ruta_pdf.stem.strip() or "carga_pdf"
    ruta_facturas = carpeta_salida / f"{nombre_base}_FACTURA_{fecha_archivo}.xlsx"
    ruta_notas_credito = carpeta_salida / f"{nombre_base}_NOTA_CREDITO_{fecha_archivo}.xlsx"
    return ruta_facturas, ruta_notas_credito


def generar_excel_carga(documentos: list[DocumentoPDF], ruta_salida: str | Path, titulo_hoja: str = "Plantilla Facturas") -> Path:
    """
    Genera un XLSX desde cero.

    Columnas:
    A: Nro_Referencia -> documento con guiones
    B: Valor original -> vacío
    C: Saldo de la factura -> vacío
    D: A pagar -> Imp. Factura sin separador de miles ni signo negativo para NCR

    Todas las celdas quedan con formato General.
    """
    ruta_salida = Path(ruta_salida)
    if ruta_salida.suffix.lower() != ".xlsx":
        ruta_salida = ruta_salida.with_suffix(".xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = titulo_hoja[:31]

    # Encabezados
    for col_idx, encabezado in enumerate(ENCABEZADOS, start=1):
        celda = ws.cell(row=1, column=col_idx, value=encabezado)
        celda.font = Font(bold=True, color="FFFFFFFF")
        celda.fill = PatternFill(fill_type="solid", start_color="FF366092", end_color="FF366092")
        celda.alignment = Alignment(horizontal="center", vertical="center")
        celda.number_format = "General"

    # Datos desde fila 2
    for fila_idx, doc in enumerate(documentos, start=2):
        ws.cell(row=fila_idx, column=1, value=doc.nro_formateado)
        ws.cell(row=fila_idx, column=2, value=None)
        ws.cell(row=fila_idx, column=3, value=None)
        ws.cell(row=fila_idx, column=4, value=doc.imp_factura)

    thin = Side(style="thin", color="FF366092")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    max_row = max(ws.max_row, 1)

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            cell.number_format = "General"

    widths = [22, 18, 20, 15]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"
    wb.save(ruta_salida)
    return ruta_salida


def contar_filas_excel(ruta_excel: str | Path) -> int:
    """Lee el XLSX generado y cuenta filas con Nro_Referencia en columna A desde la fila 2."""
    ruta_excel = Path(ruta_excel)

    if not ruta_excel.exists():
        raise FileNotFoundError(f"No existe el archivo Excel para validar: {ruta_excel}")

    wb = load_workbook(ruta_excel, data_only=True, read_only=True)
    ws = wb.active
    total = 0

    for fila in range(2, ws.max_row + 1):
        valor = ws.cell(row=fila, column=1).value
        if valor is not None and str(valor).strip():
            total += 1

    wb.close()
    return total


def procesar_pdf_a_excels_separados(ruta_pdf: str | Path, carpeta_salida: str | Path) -> ResultadoProceso:
    """
    Flujo completo:
    1. Lee PDF.
    2. Separa facturas y notas de crédito.
    3. Genera dos XLSX independientes con fecha YYYYMMDD.
    4. Lee exactamente esos XLSX generados para validar cantidad de filas.
    """
    fecha_archivo = obtener_fecha_archivo()
    documentos = extraer_documentos_pdf(ruta_pdf)
    facturas, notas_credito = separar_documentos(documentos)
    ruta_facturas, ruta_notas_credito = crear_rutas_salida(ruta_pdf, carpeta_salida, fecha_archivo)

    generar_excel_carga(facturas, ruta_facturas, titulo_hoja="Plantilla Facturas")
    generar_excel_carga(notas_credito, ruta_notas_credito, titulo_hoja="Plantilla NC")

    total_facturas_excel = contar_filas_excel(ruta_facturas)
    total_notas_credito_excel = contar_filas_excel(ruta_notas_credito)

    return ResultadoProceso(
        ruta_facturas=ruta_facturas,
        ruta_notas_credito=ruta_notas_credito,
        documentos_pdf=documentos,
        total_facturas_pdf=len(facturas),
        total_notas_credito_pdf=len(notas_credito),
        total_facturas_excel=total_facturas_excel,
        total_notas_credito_excel=total_notas_credito_excel,
        fecha_archivo=fecha_archivo,
    )


if __name__ == "__main__":
    pdf = Path("London Import S.A.12-12.pdf")
    salida = Path(".")
    resultado = procesar_pdf_a_excels_separados(pdf, salida)
    print(f"Fecha archivo: {resultado.fecha_archivo}")
    print(f"Facturas PDF: {resultado.total_facturas_pdf}")
    print(f"Notas crédito PDF: {resultado.total_notas_credito_pdf}")
    print(f"Facturas Excel: {resultado.total_facturas_excel}")
    print(f"Notas crédito Excel: {resultado.total_notas_credito_excel}")
    print(f"Excel facturas: {resultado.ruta_facturas}")
    print(f"Excel notas crédito: {resultado.ruta_notas_credito}")
